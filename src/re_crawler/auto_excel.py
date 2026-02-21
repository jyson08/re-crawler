from __future__ import annotations

import argparse
import csv
import logging
import math
import random
import re
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import requests
from openpyxl.styles import Border, Font, PatternFill, Side

LOGGER = logging.getLogger(__name__)
DELAY_MIN_SEC = 0.1
DELAY_MAX_SEC = 0.3

COL_CITY = "\uc2dc"
COL_GU = "\uad6c"
COL_DONG = "\ub3d9"
COL_COMPLEX = "\ub2e8\uc9c0"
COL_ASSIGNED_ELEM = "\ubc30\uc815\ucd08\ub4f1\ud559\uad50"
COL_ELEM_DISTANCE = "\ucd08\ub4f1\ud559\uad50\uae4c\uc9c0\uac70\ub9ac(m)"
COL_BUILT = "\uc900\uacf5\uc5f0\ub3c4(\uc5f0\uc2dd)"
COL_AGE = "\ub098\uc774"
COL_TOTAL_HOUSEHOLDS = "\uc804\uccb4\uc138\ub300\uc218"
COL_PARKING = "\uc8fc\ucc28\ub300\uc218"
COL_SUPPLY = "\uacf5\uae09\uba74\uc801"
COL_TYPE_HOUSEHOLDS = "\ud0c0\uc785\uc138\ub300\uc218"
COL_PYUNG = "\ud3c9\ud615"
COL_EXCLUSIVE = "\uc804\uc6a9\uba74\uc801"
COL_HALL_TYPE = "\ud604\uad00\uad6c\uc870"
COL_FAR = "\uc6a9\uc801\ub960"
COL_ROOM = "\ubc29\uac2f\uc218"
COL_BATH = "\ud654\uc7a5\uc2e4\uac2f\uc218"
COL_RECENT_SALE = "\ucd5c\uadfc\uc2e4\uac70\ub798\uac00(\ub9e4\ub9e4)"
COL_RECENT_SALE_DATE_FLOOR = "\ub0a0\uc9dc/\uce35"
COL_SALE = "KB\uc2dc\uc138"
COL_MIN_ASK_SALE = "\ucd5c\uc800\ub9e4\ubb3c\uac00(\ub9e4\ub9e4)"
COL_UNDERVALUE_RATIO = "\uc800\ud3c9\uac00%"
COL_RECENT_LEASE = "\ucd5c\uadfc\uc2e4\uac70\ub798\uac00(\uc804\uc138)"
COL_LEASE = "KB\uc804\uc138"
COL_LEASE_RATIO = "\uc804\uc138\uac00\uc728"
COL_LISTING_SALE = "\ub9e4\ub9e4"
COL_LISTING_LEASE = "\uc804\uc138"
COL_LISTING_MONTHLY = "\uc6d4\uc138"
COL_LINK = "\ud574\ub2f9\ub2e8\uc9c0\ub9c1\ud06c"

OUTPUT_COLUMNS = [
    COL_CITY,
    COL_GU,
    COL_DONG,
    COL_COMPLEX,
    COL_BUILT,
    COL_AGE,
    COL_TOTAL_HOUSEHOLDS,
    COL_PARKING,
    COL_HALL_TYPE,
    COL_FAR,
    COL_ASSIGNED_ELEM,
    COL_ELEM_DISTANCE,
    COL_SUPPLY,
    COL_PYUNG,
    COL_EXCLUSIVE,
    COL_TYPE_HOUSEHOLDS,
    COL_ROOM,
    COL_BATH,
    COL_RECENT_SALE,
    COL_RECENT_SALE_DATE_FLOOR,
    COL_SALE,
    COL_MIN_ASK_SALE,
    COL_UNDERVALUE_RATIO,
    COL_RECENT_LEASE,
    COL_LEASE,
    COL_LEASE_RATIO,
    COL_LISTING_SALE,
    COL_LISTING_LEASE,
    COL_LISTING_MONTHLY,
    COL_LINK,
]

KB_API = "https://api.kbland.kr"
KB_COMPLEX_LIST_URL = f"{KB_API}/land-price/price/fastPriceComplexName"
PARAM_COMPLEX_NO = "%EB%8B%A8%EC%A7%80%EA%B8%B0%EB%B3%B8%EC%9D%BC%EB%A0%A8%EB%B2%88%ED%98%B8"


@dataclass
class KbComplexCandidate:
    complex_id: int
    name: str
    address: str | None
    score: float


@dataclass
class QueryMetric:
    query: str
    started_at: datetime
    finished_at: datetime
    elapsed_sec: float
    seed_complex_id: int | None
    seed_complex_name: str | None
    nearby_candidates: int
    target_candidates: int
    attempted_complexes: int
    success_complexes: int
    failed_complexes: int

    @property
    def failure_rate_pct(self) -> float:
        if self.attempted_complexes <= 0:
            return 0.0
        return round((self.failed_complexes / self.attempted_complexes) * 100.0, 1)


def setup_logging(level: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    )


def _delay() -> None:
    if DELAY_MAX_SEC <= 0:
        return
    low = max(0.0, DELAY_MIN_SEC)
    high = max(low, DELAY_MAX_SEC)
    time.sleep(random.uniform(low, high))


def set_delay_range(min_sec: float, max_sec: float) -> None:
    global DELAY_MIN_SEC, DELAY_MAX_SEC
    DELAY_MIN_SEC = max(0.0, float(min_sec))
    DELAY_MAX_SEC = max(DELAY_MIN_SEC, float(max_sec))


def _normalize_text(text: str) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", text.lower())


def _similarity_score(query: str, target: str) -> float:
    q = _normalize_text(query)
    t = _normalize_text(target)
    if not q or not t:
        return 0.0
    if q == t:
        return 100.0
    if q in t:
        return 90.0 + min(9.0, (len(q) / max(len(t), 1)) * 10.0)
    # Loose overlap score
    common = sum(1 for ch in set(q) if ch in set(t))
    return (common / max(len(set(q)), 1)) * 70.0


def _normalize_complex_name(text: str) -> str:
    n = _normalize_text(text)
    for suffix in ("아파트", "apt"):
        n = n.replace(suffix, "")
    return n


def split_queries(raw_query: str) -> list[str]:
    parts: list[str] = []
    buf: list[str] = []
    depth = 0
    pairs = {"(": ")", "[": "]", "{": "}"}
    closers = set(pairs.values())

    for ch in raw_query:
        if ch in pairs:
            depth += 1
            buf.append(ch)
            continue
        if ch in closers:
            depth = max(0, depth - 1)
            buf.append(ch)
            continue
        if ch == "," and depth == 0:
            token = "".join(buf).strip()
            if token:
                parts.append(token)
            buf = []
            continue
        buf.append(ch)

    token = "".join(buf).strip()
    if token:
        parts.append(token)
    return parts


def _prompt_candidate_selection(query: str, candidates: list[KbComplexCandidate]) -> KbComplexCandidate:
    print(f"\n[{query}] 후보를 선택하세요 (기본값: 1)")
    for i, c in enumerate(candidates, start=1):
        addr = c.address or "-"
        print(f"{i}. {c.name} | {addr} | score={c.score:.1f} | id={c.complex_id}")

    while True:
        try:
            raw = input("번호 입력(1-10, 엔터=1): ").strip()
        except EOFError:
            LOGGER.warning("No interactive input available. Fallback to first candidate for query=%s", query)
            return candidates[0]
        if raw == "":
            return candidates[0]
        try:
            idx = int(raw)
        except ValueError:
            print("숫자를 입력해 주세요.")
            continue
        if 1 <= idx <= len(candidates):
            return candidates[idx - 1]
        print(f"1~{len(candidates)} 사이 번호를 입력해 주세요.")


def select_best_candidate(query: str, candidates: list[KbComplexCandidate], auto_score_threshold: float = 85.0) -> KbComplexCandidate:
    if not candidates:
        raise ValueError("candidates is empty")

    q_norm = _normalize_text(query)
    q_simple = _normalize_complex_name(query)
    enriched: list[tuple[KbComplexCandidate, str, str]] = []
    for c in candidates:
        n_norm = _normalize_text(c.name)
        n_simple = _normalize_complex_name(c.name)
        enriched.append((c, n_norm, n_simple))

    exact = [c for c, n_norm, n_simple in enriched if n_norm == q_norm or n_simple == q_simple]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        exact.sort(key=lambda x: x.score, reverse=True)
        return exact[0]

    contains = [c for c, _, n_simple in enriched if q_simple and (q_simple in n_simple or n_simple in q_simple)]
    if len(contains) == 1:
        return contains[0]
    if len(contains) > 1:
        contains.sort(key=lambda x: x.score, reverse=True)
        top = contains[0]
        second = contains[1]
        if top.score >= auto_score_threshold and (top.score - second.score) >= 8.0:
            return top
        return _prompt_candidate_selection(query, contains[:10])

    ranked = sorted(candidates, key=lambda x: x.score, reverse=True)
    if len(ranked) == 1:
        return ranked[0]
    if ranked[0].score >= auto_score_threshold and (ranked[0].score - ranked[1].score) >= 8.0:
        return ranked[0]
    return _prompt_candidate_selection(query, ranked[:10])


def _request_json_with_retry(session: requests.Session, url: str, retries: int = 2) -> dict[str, Any] | None:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code != 200:
                LOGGER.warning("API status=%s url=%s", resp.status_code, url)
                _delay()
                continue
            return resp.json()
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            LOGGER.warning("API request failed attempt=%d url=%s err=%s", attempt + 1, url, exc)
            _delay()
    if last_error:
        LOGGER.warning("API failed url=%s err=%s", url, last_error)
    return None


def fetch_kb_complex_index(session: requests.Session) -> list[dict[str, Any]]:
    payload = _request_json_with_retry(session, KB_COMPLEX_LIST_URL, retries=2)
    if not payload:
        return []
    items = payload.get("dataBody", {}).get("data", [])
    if not isinstance(items, list):
        return []
    return [item for item in items if isinstance(item, dict)]


def fetch_kb_complex_candidates(
    session: requests.Session,
    query: str,
    top_n: int = 10,
    index_items: list[dict[str, Any]] | None = None,
    preferred_dong: str | None = None,
) -> list[KbComplexCandidate]:
    items = index_items if index_items is not None else fetch_kb_complex_index(session)
    if not items:
        return []

    candidates: list[KbComplexCandidate] = []
    for item in items:
        name = str(item.get("단지명") or "").strip()
        if not name:
            continue
        address = item.get("주소")
        complex_id = item.get("단지기본일련번호")
        if not isinstance(complex_id, int):
            continue

        score = _similarity_score(query, name)
        if preferred_dong:
            addr_text = str(address or "")
            if preferred_dong in addr_text:
                score += 12.0
            elif score < 60:
                # 동 정보가 불일치하고 유사도도 낮으면 제외
                continue
        if score < 15:
            continue
        candidates.append(
            KbComplexCandidate(
                complex_id=complex_id,
                name=name,
                address=str(address) if address else None,
                score=score,
            )
        )

    candidates.sort(key=lambda x: x.score, reverse=True)
    return candidates[:top_n]


def _kb_endpoint(path: str, complex_id: int, extra_qs: str = "") -> str:
    base = f"{KB_API}{path}?{PARAM_COMPLEX_NO}={complex_id}"
    if extra_qs:
        return f"{base}&{extra_qs}"
    return base


def fetch_kb_complex_payloads(session: requests.Session, complex_id: int) -> dict[str, Any]:
    endpoints = {
        "info": _kb_endpoint("/land-complex/complex/info", complex_id),
        "main": _kb_endpoint("/land-complex/complex/main", complex_id),
        "typ_info": _kb_endpoint("/land-complex/complex/typInfo", complex_id),
        "mpri_by_type": _kb_endpoint("/land-complex/complex/mpriByType", complex_id),
        "school_elem": _kb_endpoint(
            "/land-complex/complexSchool/v2/list",
            complex_id,
            extra_qs="%ED%95%99%EA%B5%90%EA%B3%BC%EC%A0%95%EB%B6%84%EB%A5%98%EA%B5%AC%EB%B6%84=03",
        ),
    }

    out: dict[str, Any] = {}
    for key, url in endpoints.items():
        payload = _request_json_with_retry(session, url, retries=2)
        if not payload:
            out[key] = None
            continue
        out[key] = payload.get("dataBody", {}).get("data")
        _delay()

    # Pull per-area recent transaction snapshot (sale/lease) from BasePrcInfoNew.
    area_key = "\uba74\uc801\uc77c\ub828\ubc88\ud638"
    mpri_rows = out.get("mpri_by_type") if isinstance(out.get("mpri_by_type"), list) else []
    base_price_by_area: dict[int, dict[str, Any]] = {}
    recent_deals_by_area: dict[int, list[dict[str, Any]]] = {}
    pre_sale_chart_by_area: dict[int, dict[str, Any]] = {}
    today = date.today()
    end_ymd = today.strftime("%Y%m%d")
    start_ymd = f"{today.year - 5}{today.strftime('%m%d')}"
    for row in mpri_rows:
        if not isinstance(row, dict):
            continue
        area_id = _to_int(row.get(area_key))
        if area_id is None:
            continue
        extra_qs = f"%EB%A9%B4%EC%A0%81%EC%9D%BC%EB%A0%A8%EB%B2%88%ED%98%B8={area_id}"
        url = _kb_endpoint("/land-price/price/BasePrcInfoNew", complex_id, extra_qs=extra_qs)
        payload = _request_json_with_retry(session, url, retries=2)
        if not payload:
            continue
        data = payload.get("dataBody", {}).get("data")
        if isinstance(data, dict):
            base_price_by_area[area_id] = data
        _delay()
        # Pull detailed recent transaction list to support filtering (e.g. exclude 1st floor).
        deal_params = {
            "\ub2e8\uc9c0\uae30\ubcf8\uc77c\ub828\ubc88\ud638": complex_id,
            "\uccab\ud398\uc774\uc9c0\uac2f\uc218": 30,
            "\ud398\uc774\uc9c0\uac2f\uc218": 30,
            "\ud604\uc7ac\ud398\uc774\uc9c0": 1,
            "\uac70\ub798\uad6c\ubd84": 3,
            "\uba74\uc801\uadf8\ub8f9\uc5ec\ubd80": 0,
            "\uba74\uc801\uc77c\ub828\ubc88\ud638": area_id,
        }
        deal_url = requests.Request(
            "GET",
            f"{KB_API}/land-price/price/complex/preSalePrices",
            params=deal_params,
        ).prepare().url or ""
        deal_payload = _request_json_with_retry(session, deal_url, retries=2)
        if deal_payload:
            deal_data = deal_payload.get("dataBody", {}).get("data")
            if isinstance(deal_data, dict):
                rows = deal_data.get("dataList")
                if isinstance(rows, list):
                    recent_deals_by_area[area_id] = [r for r in rows if isinstance(r, dict)]
        _delay()
        chart_params = {
            "\uac70\ub798\uad6c\ubd84": 0,
            "\ub2e8\uc9c0\uae30\ubcf8\uc77c\ub828\ubc88\ud638": complex_id,
            "\uc870\ud68c\uad6c\ubd84": 2,
            "\uba74\uc801\uc77c\ub828\ubc88\ud638": area_id,
            "\uba74\uc801\uadf8\ub8f9\uc5ec\ubd80": 0,
            "\uc870\ud68c\uc2dc\uc791\uc77c": start_ymd,
            "\uc870\ud68c\uc885\ub8cc\uc77c": end_ymd,
        }
        chart_url = requests.Request(
            "GET",
            f"{KB_API}/land-price/price/complex/preSaleChart",
            params=chart_params,
        ).prepare().url or ""
        chart_payload = _request_json_with_retry(session, chart_url, retries=2)
        if chart_payload:
            chart_data = chart_payload.get("dataBody", {}).get("data")
            if isinstance(chart_data, dict):
                pre_sale_chart_by_area[area_id] = chart_data
        _delay()
    out["base_price_by_area"] = base_price_by_area
    out["recent_deals_by_area"] = recent_deals_by_area
    out["pre_sale_chart_by_area"] = pre_sale_chart_by_area
    return out


def fetch_kb_complex_main(session: requests.Session, complex_id: int) -> dict[str, Any] | None:
    payload = _request_json_with_retry(session, _kb_endpoint("/land-complex/complex/main", complex_id), retries=4)
    if not payload:
        return None
    data = payload.get("dataBody", {}).get("data")
    if not isinstance(data, dict):
        return None
    _delay()
    return data


def get_main_with_fallback(
    session: requests.Session,
    complex_id: int,
    main_cache: dict[int, dict[str, Any]] | None = None,
    payload_cache: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if main_cache is not None and complex_id in main_cache:
        return main_cache[complex_id]

    main_data = fetch_kb_complex_main(session, complex_id) or {}
    if not main_data and payload_cache is not None and complex_id in payload_cache:
        cached_payload = payload_cache.get(complex_id) or {}
        main_data = cached_payload.get("main") if isinstance(cached_payload.get("main"), dict) else {}
    if not main_data:
        payloads = fetch_kb_complex_payloads(session, complex_id)
        if payload_cache is not None:
            payload_cache[complex_id] = payloads
        main_data = payloads.get("main") if isinstance(payloads.get("main"), dict) else {}
    if main_cache is not None and main_data:
        main_cache[complex_id] = main_data
    return main_data or {}


def _haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    r = 6371000.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def _extract_sido_sigungu(addr: str | None) -> tuple[str | None, str | None]:
    if not addr:
        return None, None
    parts = re.sub(r"\s+", " ", addr).strip().split(" ")
    if not parts:
        return None, None
    sido = parts[0]
    sigungu = None

    if len(parts) >= 3 and (parts[1].endswith("구") or parts[1].endswith("군")):
        sigungu = parts[1]
    elif len(parts) >= 4 and parts[1].endswith("시") and (parts[2].endswith("구") or parts[2].endswith("군")):
        sigungu = f"{parts[1]} {parts[2]}"
    elif len(parts) >= 2 and parts[1].endswith("시"):
        sigungu = parts[1]

    return sido, sigungu


def fetch_kb_dong_rows_in_region(
    session: requests.Session,
    sido_name: str,
    sigungu_name: str | None,
) -> list[dict[str, Any]]:
    params: dict[str, str] = {"\uc2dc\ub3c4\uba85": sido_name}
    if sigungu_name:
        params["\uc2dc\uad70\uad6c\uba85"] = sigungu_name
    url = "https://api.kbland.kr/land-complex/map/stutDongAreaNameList"
    payload = _request_json_with_retry(
        session,
        requests.Request("GET", url, params=params).prepare().url or url,
        retries=2,
    )
    if not payload:
        return []
    rows = payload.get("dataBody", {}).get("data", [])
    if not isinstance(rows, list):
        return []
    out_rows: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        code = str(row.get("법정동코드") or "").strip()
        lat = _to_float(row.get("wgs84중심위도"))
        lng = _to_float(row.get("wgs84중심경도"))
        if not code:
            continue
        out_rows.append(
            {
                "code": code,
                "name": str(row.get("법정동명") or "").strip(),
                "lat": lat,
                "lng": lng,
            }
        )
    return out_rows


def _collect_dong_codes_from_index(
    index_items: list[dict[str, Any]],
    sido_name: str | None,
    sigungu_name: str | None,
) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for item in index_items:
        if not isinstance(item, dict):
            continue
        code = str(item.get("\ubc95\uc815\ub3d9\ucf54\ub4dc") or "").strip()
        addr = str(item.get("\uc8fc\uc18c") or "").strip()
        if not code or not addr:
            continue
        if sido_name and not addr.startswith(sido_name):
            continue
        if sigungu_name and sigungu_name not in addr:
            continue
        if code in seen:
            continue
        seen.add(code)
        codes.append(code)
    return codes


def fetch_kb_nearby_apartment_candidates(
    session: requests.Session,
    seed_candidate: KbComplexCandidate,
    seed_main: dict[str, Any],
    radius_m: float = 500.0,
    max_dong_codes: int | None = None,
    adjacent_dong_extra_m: float = 0.0,
    index_items: list[dict[str, Any]] | None = None,
) -> list[KbComplexCandidate]:
    dong_code = str(seed_main.get("법정동코드") or "").strip()
    seed_lat = _to_float(seed_main.get("wgs84위도"))
    seed_lng = _to_float(seed_main.get("wgs84경도"))
    # Fallback: if seed coords are missing, resolve from same-dong hscm list.
    if (seed_lat is None or seed_lng is None) and dong_code:
        seed_url = (
            f"{KB_API}/land-complex/complexComm/hscmList"
            f"?%EB%B2%95%EC%A0%95%EB%8F%99%EC%BD%94%EB%93%9C={dong_code}"
        )
        seed_payload = _request_json_with_retry(session, seed_url, retries=2)
        seed_rows = seed_payload.get("dataBody", {}).get("data", []) if isinstance(seed_payload, dict) else []
        if isinstance(seed_rows, list):
            for row in seed_rows:
                if not isinstance(row, dict):
                    continue
                cid = _to_int(row.get("단지기본일련번호"))
                if cid != seed_candidate.complex_id:
                    continue
                seed_lat = _to_float(row.get("wgs84위도"))
                seed_lng = _to_float(row.get("wgs84경도"))
                if seed_lat is not None and seed_lng is not None:
                    break
        _delay()
    if seed_lat is None or seed_lng is None or not dong_code:
        LOGGER.warning(
            "Nearby expansion skipped. missing lat/lng/dong_code for id=%s",
            seed_candidate.complex_id,
        )
        return [seed_candidate]

    seed_addr = (
        str(seed_main.get("구주소") or "").strip()
        or str(seed_main.get("신주소") or "").strip()
        or str(seed_main.get("도로기본주소") or "").strip()
    )
    sido_name, sigungu_name = _extract_sido_sigungu(seed_addr)
    dong_rows: list[dict[str, Any]] = []
    if sido_name:
        dong_rows = fetch_kb_dong_rows_in_region(
            session=session,
            sido_name=sido_name,
            sigungu_name=sigungu_name,
        )
    dong_codes: list[str] = []
    if dong_rows:
        # 정확도 우선: 시군구 내 동코드를 모두 탐색하고 실제 단지 좌표 거리로만 필터링
        dong_codes = [str(r.get("code") or "").strip() for r in dong_rows if str(r.get("code") or "").strip()]
    if not dong_codes and index_items:
        dong_codes = _collect_dong_codes_from_index(
            index_items=index_items,
            sido_name=sido_name,
            sigungu_name=sigungu_name,
        )
    if not dong_codes:
        dong_codes = [dong_code]
    if dong_code not in dong_codes:
        dong_codes.insert(0, dong_code)
    if max_dong_codes is not None and max_dong_codes > 0 and len(dong_codes) > max_dong_codes:
        # 속도 우선 모드에서만 제한
        dedup = []
        seen = set()
        for c in dong_codes:
            if c in seen:
                continue
            seen.add(c)
            dedup.append(c)
        dong_codes = dedup[:max_dong_codes]

    scored: list[tuple[float, KbComplexCandidate]] = []
    for code in dong_codes:
        url = (
            f"{KB_API}/land-complex/complexComm/hscmList"
            f"?%EB%B2%95%EC%A0%95%EB%8F%99%EC%BD%94%EB%93%9C={code}"
        )
        payload = _request_json_with_retry(session, url, retries=2)
        if not payload:
            continue
        rows = payload.get("dataBody", {}).get("data", [])
        if not isinstance(rows, list):
            continue
        _delay()

        for row in rows:
            if not isinstance(row, dict):
                continue
            if str(row.get("매물종별구분") or "") != "01":
                continue

            cid = _to_int(row.get("단지기본일련번호"))
            if cid is None:
                continue

            lat = _to_float(row.get("wgs84위도"))
            lng = _to_float(row.get("wgs84경도"))
            if lat is None or lng is None:
                continue

            distance = _haversine_m(seed_lat, seed_lng, lat, lng)
            row_dong_code = str(row.get("법정동코드") or "").strip()
            limit_m = radius_m if row_dong_code == dong_code else (radius_m + max(0.0, adjacent_dong_extra_m))
            if distance > limit_m:
                continue

            name = str(row.get("단지명") or "").strip()
            if not name:
                continue

            scored.append(
                (
                    distance,
                    KbComplexCandidate(
                        complex_id=cid,
                        name=name,
                        address=None,
                        score=seed_candidate.score if cid == seed_candidate.complex_id else 0.0,
                    ),
                )
            )

    if not scored:
        return [seed_candidate]

    scored.sort(key=lambda x: (x[0], x[1].name))

    unique: dict[int, KbComplexCandidate] = {}
    for _, cand in scored:
        unique[cand.complex_id] = cand
    if seed_candidate.complex_id not in unique:
        unique[seed_candidate.complex_id] = seed_candidate
    return list(unique.values())


def _extract_city_gu_dong(addr: str | None) -> tuple[str | None, str | None, str | None]:
    if not addr:
        return None, None, None
    parts = re.sub(r"\s+", " ", addr).strip().split(" ")
    city = parts[0] if len(parts) > 0 else None
    gu = parts[1] if len(parts) > 1 else None
    dong = parts[2] if len(parts) > 2 else None
    return city, gu, dong


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", ""))
    except Exception:  # noqa: BLE001
        return None


def _round1(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 1)


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "")))
    except Exception:  # noqa: BLE001
        return None


def _calc_pyung(supply_m2: float | None) -> float | None:
    if supply_m2 is None:
        return None
    return round(supply_m2 * 0.3025, 1)


def _calc_lease_ratio(lease_price: int | None, sale_price: int | None) -> float | None:
    if lease_price is None or sale_price in (None, 0):
        return None
    return round((lease_price / sale_price) * 100, 1)


def _calc_undervalue_ratio(recent_sale_price: int | None, kb_sale_price: int | None) -> float | None:
    if recent_sale_price is None or kb_sale_price in (None, 0):
        return None
    return round((recent_sale_price / kb_sale_price) * 100, 1)


def _format_date_floor(contract_yyyymmdd: Any, floor: Any) -> str | None:
    raw = str(contract_yyyymmdd or "").strip()
    floor_text = str(floor or "").strip()
    if not raw and not floor_text:
        return None

    date_text = raw
    if len(raw) == 8 and raw.isdigit():
        date_text = f"{raw[2:4]}.{raw[4:6]}.{raw[6:8]}"

    if floor_text and not floor_text.endswith("층"):
        floor_text = f"{floor_text}층"

    if date_text and floor_text:
        return f"{date_text}/{floor_text}"
    return date_text or floor_text


def _parse_floor_number(floor: Any) -> int | None:
    text = str(floor or "").strip()
    if not text:
        return None
    nums = re.findall(r"-?\d+", text)
    if not nums:
        return None
    try:
        # Use the last number so values like "607/5층" resolve to 5.
        return int(nums[-1])
    except Exception:  # noqa: BLE001
        return None


def _pick_recent_trade_from_rows(
    rows: list[dict[str, Any]],
    trade_name: str,
) -> tuple[int | None, str | None]:
    date_key = "\uacc4\uc57d\ub144\uc6d4\uc77c"
    name_key = "\ubb3c\uac74\uac70\ub798\uba85"
    cancel_key = "\uacc4\uc57d\ucde8\uc18c\uc5ec\ubd80"
    floor_key = "\ud574\ub2f9\uce35\uc218"
    sale_price_key = "\ub9e4\ub9e4\uc2e4\uac70\ub798\uae08\uc561"
    lease_price_key = "\uc804\uc138\uc2e4\uac70\ub798\uae08\uc561"

    candidates: list[tuple[str, int | None, int]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        if str(row.get(cancel_key) or "0") == "1":
            continue
        if str(row.get(name_key) or "").strip() != trade_name:
            continue
        floor_no = _parse_floor_number(row.get(floor_key))
        if floor_no == 1:
            continue
        if trade_name == "\ub9e4\ub9e4":
            price = _to_int(row.get(sale_price_key))
        else:
            price = _to_int(row.get(lease_price_key))
        if price is None:
            continue
        raw_date = str(row.get(date_key) or "").strip()
        if not (len(raw_date) == 8 and raw_date.isdigit()):
            raw_date = "00000000"
        candidates.append((raw_date, floor_no, price))

    if not candidates:
        return None, None
    candidates.sort(key=lambda x: x[0], reverse=True)
    best_date, best_floor, best_price = candidates[0]
    floor_text = f"{best_floor}\uce35" if best_floor is not None else None
    return best_price, _format_date_floor(best_date, floor_text)


def build_dataframe_from_kb(query: str, candidate: KbComplexCandidate, payloads: dict[str, Any]) -> pd.DataFrame:
    info = payloads.get("info") if isinstance(payloads.get("info"), dict) else {}
    main = payloads.get("main") if isinstance(payloads.get("main"), dict) else {}
    typ_rows = payloads.get("typ_info") if isinstance(payloads.get("typ_info"), list) else []
    mpri_rows = payloads.get("mpri_by_type") if isinstance(payloads.get("mpri_by_type"), list) else []
    base_price_by_area = payloads.get("base_price_by_area") if isinstance(payloads.get("base_price_by_area"), dict) else {}
    recent_deals_by_area = payloads.get("recent_deals_by_area") if isinstance(payloads.get("recent_deals_by_area"), dict) else {}
    pre_sale_chart_by_area = payloads.get("pre_sale_chart_by_area") if isinstance(payloads.get("pre_sale_chart_by_area"), dict) else {}
    school_rows = payloads.get("school_elem") if isinstance(payloads.get("school_elem"), list) else []

    complex_name = str(main.get("단지명") or info.get("단지명") or candidate.name or query)
    completion_year = _to_int(main.get("준공년"))
    age = date.today().year - completion_year if completion_year else None
    total_households = _to_int(main.get("총세대수"))
    parking_ratio = _to_float(main.get("세대당주차대수비율"))
    if parking_ratio is None:
        total_parking = _to_int(main.get("총주차대수"))
        if total_parking is not None and total_households not in (None, 0):
            parking_ratio = total_parking / total_households
    parking_text = f"{round(parking_ratio, 2):.2f}대" if parking_ratio is not None else None
    floor_area_ratio = _to_float(main.get("용적률내용"))
    if floor_area_ratio is None:
        floor_area_ratio = _to_float(main.get("용적률"))
    address = (
        main.get("구주소")
        or main.get("신주소")
        or main.get("도로기본주소")
        or candidate.address
    )
    city, gu, dong = _extract_city_gu_dong(str(address) if address else None)

    assigned_school_name = None
    assigned_school_distance = None
    if school_rows:
        def _school_key(s: dict[str, Any]) -> tuple[int, float]:
            assigned = 0 if str(s.get("\ubc30\uc815\uc5ec\ubd80") or "") == "1" else 1
            dist = _to_float(s.get("\uac70\ub9ac"))
            return assigned, dist if dist is not None else float("inf")

        best_school = sorted([s for s in school_rows if isinstance(s, dict)], key=_school_key)[0]
        assigned_school_name = best_school.get("\ud559\uad50\uba85")
        assigned_school_distance = _round1(_to_float(best_school.get("\uac70\ub9ac")))

    typ_map: dict[int, dict[str, Any]] = {}
    for t in typ_rows:
        if not isinstance(t, dict):
            continue
        area_id = _to_int(t.get("면적일련번호"))
        if area_id is None:
            continue
        typ_map[area_id] = t

    rows: list[dict[str, Any]] = []
    for m in mpri_rows:
        if not isinstance(m, dict):
            continue

        area_id = _to_int(m.get("면적일련번호"))
        typ = typ_map.get(area_id, {})
        base_data = base_price_by_area.get(area_id, {})
        chart_data = pre_sale_chart_by_area.get(area_id, {})

        supply_m2 = _round1(_to_float(m.get("공급면적")))
        exclusive_m2 = _round1(_to_float(m.get("전용면적")))
        if exclusive_m2 is not None and exclusive_m2 <= 59:
            continue
        if exclusive_m2 is not None and exclusive_m2 > 85:
            continue

        sale = _to_int(m.get("매매일반거래가"))
        min_ask_sale = _to_int(m.get("매매하한가"))
        chart_min_ask_sale = _to_int(chart_data.get("\ub9e4\ubb3c\ub9e4\ub9e4\ucd5c\uc800\uac00")) if isinstance(chart_data, dict) else None
        if chart_min_ask_sale is not None:
            min_ask_sale = chart_min_ask_sale
        if sale is None:
            low = min_ask_sale
            high = _to_int(m.get("매매상한가"))
            if low is not None and high is not None:
                sale = int(round((low + high) / 2))
            else:
                sale = low or high

        lease = _to_int(m.get("전세일반거래가"))

        recent_sale_price = None
        recent_sale_date_floor = None
        recent_lease_price = None
        recent_rows = recent_deals_by_area.get(area_id, []) if area_id is not None else []
        if isinstance(recent_rows, list) and recent_rows:
            recent_sale_price, recent_sale_date_floor = _pick_recent_trade_from_rows(
                recent_rows,
                trade_name="\ub9e4\ub9e4",
            )
            recent_lease_price, _ = _pick_recent_trade_from_rows(
                recent_rows,
                trade_name="\uc804\uc138",
            )

        # Fallback to base snapshot when detailed transaction rows are missing.
        if isinstance(base_data, dict):
            if recent_sale_price is None:
                recent_obj = base_data.get("\ucd5c\uadfc\uc2e4\uac70\ub798\uac00")
                if isinstance(recent_obj, dict):
                    recent_sale_floor = _parse_floor_number(recent_obj.get("\uac70\ub798\uce35"))
                    if recent_sale_floor != 1:
                        recent_sale_price = _to_int(recent_obj.get("\uac70\ub798\uae08\uc561"))
                        recent_sale_date_floor = _format_date_floor(
                            recent_obj.get("\uacc4\uc57d\ub144\uc6d4\uc77c"),
                            recent_sale_floor,
                        )
            if recent_lease_price is None:
                sise_rows = base_data.get("\uc2dc\uc138")
                if isinstance(sise_rows, list) and sise_rows and isinstance(sise_rows[0], dict):
                    recent_lease_floor = _parse_floor_number(sise_rows[0].get("\uc804\uc138\ud574\ub2f9\uce35\uc218"))
                    if recent_lease_floor != 1:
                        recent_lease_price = _to_int(sise_rows[0].get("\uc804\uc138\uac70\ub798\uae08\uc561"))

        sale_count = _to_int(m.get("\ub9e4\ub9e4\uac74\uc218")) or 0
        lease_count = _to_int(m.get("\uc804\uc138\uac74\uc218")) or 0
        monthly_count = _to_int(m.get("\uc6d4\uc138\uac74\uc218")) or 0

        rows.append(
            {
                COL_CITY: city,
                COL_GU: gu,
                COL_DONG: dong,
                COL_COMPLEX: complex_name,
                COL_ASSIGNED_ELEM: assigned_school_name,
                COL_ELEM_DISTANCE: assigned_school_distance,
                COL_BUILT: completion_year,
                COL_AGE: age,
                COL_TOTAL_HOUSEHOLDS: total_households,
                COL_PARKING: parking_text,
                COL_FAR: floor_area_ratio,
                COL_SUPPLY: supply_m2,
                COL_TYPE_HOUSEHOLDS: _to_int(typ.get("세대수")) or _to_int(m.get("세대수")),
                COL_PYUNG: _calc_pyung(supply_m2),
                COL_EXCLUSIVE: exclusive_m2,
                COL_HALL_TYPE: main.get("\ud604\uad00\uad6c\uc870"),
                COL_ROOM: _to_int(typ.get("방수")),
                COL_BATH: _to_int(typ.get("욕실수")),
                COL_RECENT_SALE: recent_sale_price,
                COL_RECENT_SALE_DATE_FLOOR: recent_sale_date_floor,
                COL_SALE: sale,
                COL_MIN_ASK_SALE: min_ask_sale,
                COL_UNDERVALUE_RATIO: _calc_undervalue_ratio(min_ask_sale, sale),
                COL_RECENT_LEASE: recent_lease_price,
                COL_LEASE: lease,
                COL_LEASE_RATIO: _calc_lease_ratio(lease, sale),
                COL_LISTING_SALE: sale_count,
                COL_LISTING_LEASE: lease_count,
                COL_LISTING_MONTHLY: monthly_count,
                COL_LINK: f"https://kbland.kr/c/{candidate.complex_id}",
            }
        )

    if not rows:
        LOGGER.info(
            "Skip complex by area filter. id=%s name=%s (no rows in exclusive area range)",
            candidate.complex_id,
            complex_name,
        )
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def save_output(df: pd.DataFrame, query: str, output_dir: str = "./output") -> Path:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    safe_name = "".join(ch for ch in query if ch not in '\\/:*?"<>|\r\n\t').strip() or "complex"
    today = date.today().strftime("%Y%m%d")
    out_path = Path(output_dir) / f"{safe_name}_{today}.xlsx"

    # Two-line header: top-level group + detail column.
    group_map = {
        COL_CITY: "\uc704\uce58",
        COL_GU: "\uc704\uce58",
        COL_DONG: "\uc704\uce58",
        COL_COMPLEX: "\ub2e8\uc9c0",
        COL_ASSIGNED_ELEM: "\ud559\uad70",
        COL_ELEM_DISTANCE: "\ud559\uad70",
        COL_BUILT: "\ub2e8\uc9c0",
        COL_AGE: "\ub2e8\uc9c0",
        COL_TOTAL_HOUSEHOLDS: "\ub2e8\uc9c0",
        COL_PARKING: "\ub2e8\uc9c0",
        COL_FAR: "\ub2e8\uc9c0",
        COL_SUPPLY: "\uba74\uc801/\ud0c0\uc785",
        COL_TYPE_HOUSEHOLDS: "\uba74\uc801/\ud0c0\uc785",
        COL_PYUNG: "\uba74\uc801/\ud0c0\uc785",
        COL_EXCLUSIVE: "\uba74\uc801/\ud0c0\uc785",
        COL_HALL_TYPE: "\ub2e8\uc9c0",
        COL_ROOM: "\uba74\uc801/\ud0c0\uc785",
        COL_BATH: "\uba74\uc801/\ud0c0\uc785",
        COL_RECENT_SALE: "\uac00\uaca9",
        COL_RECENT_SALE_DATE_FLOOR: "\uac00\uaca9",
        COL_SALE: "\uac00\uaca9",
        COL_MIN_ASK_SALE: "\uac00\uaca9",
        COL_UNDERVALUE_RATIO: "\uac00\uaca9",
        COL_RECENT_LEASE: "\uac00\uaca9",
        COL_LEASE: "\uac00\uaca9",
        COL_LEASE_RATIO: "\uac00\uaca9",
        COL_LISTING_SALE: "\ub9e4\ubb3c\uc218",
        COL_LISTING_LEASE: "\ub9e4\ubb3c\uc218",
        COL_LISTING_MONTHLY: "\ub9e4\ubb3c\uc218",
        COL_LINK: "\ub9c1\ud06c",
    }
    def _write_with_two_headers(path: Path) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # write body without header, starting at row 3
            df.to_excel(writer, index=False, header=False, startrow=2, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]
            header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
            header_font = Font(bold=True)
            thin_side = Side(style="thin", color="BFBFBF")
            thick_side = Side(style="thick", color="808080")

            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_idx, value=group_map.get(col_name, ""))
                ws.cell(row=2, column=col_idx, value=col_name)
                ws.cell(row=1, column=col_idx).fill = header_fill
                ws.cell(row=2, column=col_idx).fill = header_fill
                ws.cell(row=1, column=col_idx).font = header_font
                ws.cell(row=2, column=col_idx).font = header_font
                ws.cell(row=1, column=col_idx).border = Border(
                    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
                )
                ws.cell(row=2, column=col_idx).border = Border(
                    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
                )

            # Alternate fills by complex block for readability.
            complex_idx = {name: i + 1 for i, name in enumerate(df.columns)}.get(COL_COMPLEX)
            if complex_idx is not None:
                block_fills = [
                    PatternFill(fill_type="solid", fgColor="FFFFFF"),
                    PatternFill(fill_type="solid", fgColor="F2F2F2"),
                ]
                current_complex = None
                block_no = -1
                for r in range(3, ws.max_row + 1):
                    value = ws.cell(row=r, column=complex_idx).value
                    if value != current_complex:
                        block_no += 1
                        current_complex = value
                    row_fill = block_fills[block_no % len(block_fills)]
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r, column=c).fill = row_fill

            # Number formats
            money_cols = {COL_RECENT_SALE, COL_SALE, COL_MIN_ASK_SALE, COL_RECENT_LEASE, COL_LEASE}
            ratio_cols = {COL_UNDERVALUE_RATIO, COL_LEASE_RATIO}
            far_col = COL_FAR
            area_cols = {COL_SUPPLY, COL_PYUNG, COL_EXCLUSIVE}
            households_col = COL_TOTAL_HOUSEHOLDS
            col_name_to_idx = {name: i + 1 for i, name in enumerate(df.columns)}
            max_row = ws.max_row

            for name in money_cols:
                idx = col_name_to_idx.get(name)
                if idx is None:
                    continue
                for r in range(3, max_row + 1):
                    ws.cell(row=r, column=idx).number_format = "#,##0"

            for name in ratio_cols:
                ratio_idx = col_name_to_idx.get(name)
                if ratio_idx is None:
                    continue
                for r in range(3, max_row + 1):
                    ws.cell(row=r, column=ratio_idx).number_format = '0.0"%"'

            far_idx = col_name_to_idx.get(far_col)
            if far_idx is not None:
                for r in range(3, max_row + 1):
                    ws.cell(row=r, column=far_idx).number_format = '0"%"'

            for name in area_cols:
                idx = col_name_to_idx.get(name)
                if idx is None:
                    continue
                for r in range(3, max_row + 1):
                    ws.cell(row=r, column=idx).number_format = "00.0"

            households_idx = col_name_to_idx.get(households_col)
            if households_idx is not None:
                for r in range(3, max_row + 1):
                    ws.cell(row=r, column=households_idx).number_format = "#,##0"

            # Highlight rows where undervalue ratio is <= 100%.
            undervalue_idx = col_name_to_idx.get(COL_UNDERVALUE_RATIO)
            cond_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
            cond_font = Font(bold=True)
            if undervalue_idx is not None:
                for r in range(3, max_row + 1):
                    undervalue_val = _to_float(ws.cell(row=r, column=undervalue_idx).value)
                    if undervalue_val is None:
                        continue
                    if 0.0 < undervalue_val <= 100.0:
                        for c in range(1, len(df.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.fill = cond_fill
                            cell.font = cond_font

            # Thick right border on group boundaries for the whole table.
            groups = [group_map.get(c, "") for c in df.columns]
            boundary_cols: list[int] = []
            for i in range(1, len(groups)):
                if groups[i] != groups[i - 1]:
                    boundary_cols.append(i)  # 1-based previous column index
            boundary_cols.append(len(groups))

            max_row = ws.max_row
            for c in boundary_cols:
                for r in range(1, max_row + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = Border(
                        left=cell.border.left if cell.border else thin_side,
                        right=thick_side,
                        top=cell.border.top if cell.border else thin_side,
                        bottom=cell.border.bottom if cell.border else thin_side,
                    )

    try:
        _write_with_two_headers(out_path)
    except PermissionError:
        alt = Path(output_dir) / f"{safe_name}_{today}_{int(time.time())}.xlsx"
        _write_with_two_headers(alt)
        out_path = alt
    return out_path


def create_kb_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "user-agent": "Mozilla/5.0",
            "accept": "application/json, text/plain, */*",
            "referer": "https://kbland.kr/",
            "webservice": "pc",
        }
    )
    return session


def _extract_marker_row(cand: KbComplexCandidate, main_data: dict[str, Any], seed_query: str, is_seed: bool) -> dict[str, Any] | None:
    lat = _to_float(main_data.get("wgs84위도"))
    lng = _to_float(main_data.get("wgs84경도"))
    if lat is None or lng is None:
        return None
    built_year = _to_int(main_data.get("준공년"))
    households = _to_int(main_data.get("총세대수"))
    parking_ratio = _to_float(main_data.get("세대당주차대수비율"))
    if parking_ratio is None:
        total_parking = _to_int(main_data.get("총주차대수"))
        if total_parking is not None and households not in (None, 0):
            parking_ratio = total_parking / households
    parking_text = f"{round(parking_ratio, 2):.2f}대" if parking_ratio is not None else None
    hall_type = str(main_data.get("현관구조") or "").strip() or None
    return {
        "complex_id": cand.complex_id,
        "complex_name": str(main_data.get("단지명") or cand.name),
        "lat": lat,
        "lng": lng,
        "seed_query": seed_query,
        "is_seed": is_seed,
        "built_year": built_year,
        "households": households,
        "parking": parking_text,
        "hall_type": hall_type,
    }


def save_query_metrics(metrics: list[QueryMetric], output_dir: str = "./output") -> Path:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(output_dir) / "query_metrics.csv"
    file_exists = out_path.exists()
    with out_path.open("a", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        if not file_exists:
            writer.writerow(
                [
                    "run_at",
                    "query",
                    "started_at",
                    "finished_at",
                    "elapsed_sec",
                    "seed_complex_id",
                    "seed_complex_name",
                    "nearby_candidates",
                    "target_candidates",
                    "attempted_complexes",
                    "success_complexes",
                    "failed_complexes",
                    "failure_rate_pct",
                ]
            )
        run_at = datetime.now().isoformat(timespec="seconds")
        for m in metrics:
            writer.writerow(
                [
                    run_at,
                    m.query,
                    m.started_at.isoformat(timespec="seconds"),
                    m.finished_at.isoformat(timespec="seconds"),
                    m.elapsed_sec,
                    m.seed_complex_id,
                    m.seed_complex_name,
                    m.nearby_candidates,
                    m.target_candidates,
                    m.attempted_complexes,
                    m.success_complexes,
                    m.failed_complexes,
                    m.failure_rate_pct,
                ]
            )
    return out_path


def preview_candidates(
    raw_query: str,
    radius_m: float = 500.0,
    min_households: int = 290,
    preferred_dong: str | None = None,
    fast_mode: bool = True,
    max_dong_codes: int | None = None,
    index_items: list[dict[str, Any]] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[int], list[tuple[str, KbComplexCandidate]]]:
    if fast_mode:
        set_delay_range(0.05, 0.15)
    else:
        set_delay_range(0.5, 1.5)

    queries = split_queries(raw_query)
    if not queries:
        raise ValueError("단지명이 비어 있습니다.")

    session = create_kb_session()
    kb_index_items = index_items if index_items is not None else fetch_kb_complex_index(session)
    if not kb_index_items:
        raise ValueError("KB 단지 인덱스를 불러오지 못했습니다.")

    selected_info: list[tuple[str, KbComplexCandidate]] = []
    processed_ids: set[int] = set()
    preview_rows: list[dict[str, Any]] = []
    marker_rows: list[dict[str, Any]] = []

    for query in queries:
        candidates = fetch_kb_complex_candidates(
            session,
            query=query,
            top_n=10,
            index_items=kb_index_items,
            preferred_dong=preferred_dong,
        )
        if not candidates:
            continue
        seed = select_best_candidate(query, candidates)
        selected_info.append((query, seed))

        seed_main = fetch_kb_complex_main(session, seed.complex_id) or {}
        nearby = fetch_kb_nearby_apartment_candidates(
            session=session,
            seed_candidate=seed,
            seed_main=seed_main,
            radius_m=radius_m,
            max_dong_codes=max_dong_codes,
            index_items=kb_index_items,
        )
        for cand in nearby:
            if cand.complex_id in processed_ids:
                continue
            if cand.complex_id == seed.complex_id:
                main_data = seed_main
            else:
                main_data = get_main_with_fallback(session, cand.complex_id)
            households = _to_int(main_data.get("총세대수"))
            # Keep unknown-household candidates in preview to avoid false negatives
            # from transient API misses; strict filtering still applies when known.
            if households is not None and households < min_households:
                continue

            processed_ids.add(cand.complex_id)
            built_year = _to_int(main_data.get("준공년"))
            addr = str(main_data.get("구주소") or main_data.get("신주소") or main_data.get("도로기본주소") or "")
            city, gu, dong = _extract_city_gu_dong(addr if addr else None)
            parking_ratio = _to_float(main_data.get("세대당주차대수비율"))
            if parking_ratio is None:
                total_parking = _to_int(main_data.get("총주차대수"))
                if total_parking is not None and households not in (None, 0):
                    parking_ratio = total_parking / households
            parking_text = f"{round(parking_ratio, 2):.2f}대" if parking_ratio is not None else None
            hall_type = main_data.get("현관구조")

            preview_rows.append(
                {
                    "complex_id": cand.complex_id,
                    "단지명": str(main_data.get("단지명") or cand.name),
                    "시": city,
                    "구": gu,
                    "동": dong,
                    "준공연도": built_year,
                    "세대수": households,
                    "주차대수": parking_text,
                    "현관구조": hall_type,
                    "seed_query": query,
                    "is_seed": cand.complex_id == seed.complex_id,
                }
            )

            marker = _extract_marker_row(
                cand=cand,
                main_data=main_data,
                seed_query=query,
                is_seed=(cand.complex_id == seed.complex_id),
            )
            if marker:
                marker_rows.append(marker)

    if not preview_rows:
        raise ValueError("조건에 맞는 후보 단지가 없습니다.")

    preview_df = pd.DataFrame(preview_rows)
    preview_df = preview_df.sort_values(by=["is_seed", "시", "구", "동", "단지명"], ascending=[False, True, True, True, True])
    markers_df = pd.DataFrame(
        marker_rows,
        columns=[
            "complex_id",
            "complex_name",
            "lat",
            "lng",
            "seed_query",
            "is_seed",
            "built_year",
            "households",
            "parking",
            "hall_type",
        ],
    ).drop_duplicates(subset=["complex_id"], keep="first")
    candidate_ids = preview_df["complex_id"].astype(int).tolist()
    return preview_df, markers_df, candidate_ids, selected_info


def collect_dataset(
    raw_query: str,
    radius_m: float = 500.0,
    min_households: int = 290,
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
    fast_mode: bool = True,
    max_dong_codes: int | None = None,
    index_items: list[dict[str, Any]] | None = None,
    preferred_dong: str | None = None,
    candidate_ids: list[int] | None = None,
) -> tuple[pd.DataFrame, list[tuple[str, KbComplexCandidate]], list[KbComplexCandidate], pd.DataFrame, list[QueryMetric]]:
    if fast_mode:
        set_delay_range(0.05, 0.15)
    else:
        set_delay_range(0.5, 1.5)

    queries = split_queries(raw_query)
    if not queries:
        raise ValueError("단지명이 비어 있습니다.")

    session = create_kb_session()
    if progress_callback:
        progress_callback({"event": "prepare", "stage": "index_start", "message": "단지 인덱스 로딩 중..."})
    kb_index_items = index_items if index_items is not None else fetch_kb_complex_index(session)
    if not kb_index_items:
        raise ValueError("KB 단지 인덱스를 불러오지 못했습니다.")
    if progress_callback:
        progress_callback(
            {
                "event": "prepare",
                "stage": "index_done",
                "message": f"단지 인덱스 로딩 완료 ({len(kb_index_items):,}건)",
            }
        )

    all_frames: list[pd.DataFrame] = []
    selected_info: list[tuple[str, KbComplexCandidate]] = []
    crawled_info: list[KbComplexCandidate] = []
    marker_rows: list[dict[str, Any]] = []
    marker_ids: set[int] = set()
    processed_complex_ids: set[int] = set()
    metrics: list[QueryMetric] = []
    main_cache: dict[int, dict[str, Any]] = {}
    payload_cache: dict[int, dict[str, Any]] = {}

    if candidate_ids:
        chosen_ids = []
        seen = set()
        for cid in candidate_ids:
            iv = _to_int(cid)
            if iv is None or iv in seen:
                continue
            seen.add(iv)
            chosen_ids.append(iv)
        if not chosen_ids:
            raise ValueError("수집할 후보 단지 ID가 없습니다.")

        started_at = datetime.now()
        attempted = 0
        success = 0
        failed = 0
        seed_query = queries[0]

        if progress_callback:
            progress_callback(
                {
                    "event": "query_target_ready",
                    "query": seed_query,
                    "current": 0,
                    "total": len(chosen_ids),
                }
            )

        for idx, cid in enumerate(chosen_ids, start=1):
            attempted += 1
            payloads = fetch_kb_complex_payloads(session, cid)
            main_data = payloads.get("main") if isinstance(payloads.get("main"), dict) else {}
            cand = KbComplexCandidate(
                complex_id=cid,
                name=str(main_data.get("단지명") or f"complex_{cid}"),
                address=str(main_data.get("구주소") or main_data.get("신주소") or "") or None,
                score=0.0,
            )
            df = build_dataframe_from_kb(query=seed_query, candidate=cand, payloads=payloads)
            if df.empty:
                failed += 1
                if progress_callback:
                    progress_callback(
                        {
                            "event": "query_progress",
                            "query": seed_query,
                            "current": idx,
                            "total": len(chosen_ids),
                            "complex_name": cand.name,
                            "success": False,
                        }
                    )
                continue
            success += 1
            crawled_info.append(cand)
            all_frames.append(df)

            marker = _extract_marker_row(
                cand=cand,
                main_data=main_data,
                seed_query=seed_query,
                is_seed=(idx == 1),
            )
            if marker and cand.complex_id not in marker_ids:
                marker_ids.add(cand.complex_id)
                marker_rows.append(marker)

            if progress_callback:
                progress_callback(
                    {
                        "event": "query_progress",
                        "query": seed_query,
                        "current": idx,
                        "total": len(chosen_ids),
                        "complex_name": cand.name,
                        "success": True,
                    }
                )

        finished_at = datetime.now()
        metrics.append(
            QueryMetric(
                query=seed_query,
                started_at=started_at,
                finished_at=finished_at,
                elapsed_sec=round((finished_at - started_at).total_seconds(), 2),
                seed_complex_id=chosen_ids[0],
                seed_complex_name=None,
                nearby_candidates=len(chosen_ids),
                target_candidates=len(chosen_ids),
                attempted_complexes=attempted,
                success_complexes=success,
                failed_complexes=failed,
            )
        )

        if not all_frames:
            raise ValueError("수집 가능한 단지가 없습니다.")

        result_df = pd.concat(all_frames, ignore_index=True)
        markers_df = pd.DataFrame(
            marker_rows,
            columns=[
                "complex_id",
                "complex_name",
                "lat",
                "lng",
                "seed_query",
                "is_seed",
                "built_year",
                "households",
                "parking",
                "hall_type",
            ],
        )
        return result_df, selected_info, crawled_info, markers_df, metrics

    total_queries = len(queries)
    for q_idx, query in enumerate(queries, start=1):
        if progress_callback:
            progress_callback(
                {
                    "event": "prepare",
                    "stage": "query_start",
                    "message": f"쿼리 준비 중 ({q_idx}/{total_queries}): {query}",
                }
            )
        query_started_at = datetime.now()
        attempted_complexes = 0
        success_complexes = 0
        failed_complexes = 0
        nearby_candidates_count = 0
        target_candidates_count = 0
        seed_complex_id: int | None = None
        seed_complex_name: str | None = None

        candidates = fetch_kb_complex_candidates(
            session,
            query=query,
            top_n=10,
            index_items=kb_index_items,
            preferred_dong=preferred_dong,
        )
        if not candidates:
            LOGGER.warning("KB 단지 후보를 찾지 못했습니다: %s", query)
            query_finished_at = datetime.now()
            metrics.append(
                QueryMetric(
                    query=query,
                    started_at=query_started_at,
                    finished_at=query_finished_at,
                    elapsed_sec=round((query_finished_at - query_started_at).total_seconds(), 2),
                    seed_complex_id=None,
                    seed_complex_name=None,
                    nearby_candidates=0,
                    target_candidates=0,
                    attempted_complexes=0,
                    success_complexes=0,
                    failed_complexes=0,
                )
            )
            continue

        seed = select_best_candidate(query, candidates)
        seed_complex_id = seed.complex_id
        seed_complex_name = seed.name
        selected_info.append((query, seed))
        LOGGER.info("Selected seed candidate: id=%s name=%s score=%.1f", seed.complex_id, seed.name, seed.score)

        seed_payloads = payload_cache.get(seed.complex_id)
        if seed_payloads is None:
            seed_payloads = fetch_kb_complex_payloads(session, seed.complex_id)
            payload_cache[seed.complex_id] = seed_payloads
        seed_main = seed_payloads.get("main") if isinstance(seed_payloads.get("main"), dict) else {}
        if seed_main:
            main_cache[seed.complex_id] = seed_main
        nearby_candidates = fetch_kb_nearby_apartment_candidates(
            session=session,
            seed_candidate=seed,
            seed_main=seed_main,
            radius_m=radius_m,
            max_dong_codes=max_dong_codes,
            index_items=kb_index_items,
        )
        LOGGER.info(
            "Nearby apartment candidates within %.0fm from seed(id=%s): %d",
            radius_m,
            seed.complex_id,
            len(nearby_candidates),
        )
        nearby_candidates_count = len(nearby_candidates)

        target_candidates: list[tuple[KbComplexCandidate, dict[str, Any], int]] = []
        for cand in nearby_candidates:
            if cand.complex_id in processed_complex_ids:
                continue

            if cand.complex_id == seed.complex_id:
                main_data = seed_main
            else:
                main_data = main_cache.get(cand.complex_id)
                if main_data is None:
                    main_data = get_main_with_fallback(
                        session=session,
                        complex_id=cand.complex_id,
                        main_cache=main_cache,
                        payload_cache=payload_cache,
                    )

            households = _to_int(main_data.get("총세대수"))
            if households is not None and households < min_households:
                LOGGER.info(
                    "Skip complex by households threshold. id=%s name=%s households=%s threshold=%s",
                    cand.complex_id,
                    cand.name,
                    households,
                    min_households,
                )
                continue

            target_candidates.append((cand, main_data, households))

        LOGGER.info(
            "Filtered target complexes (households >= %s): %d",
            min_households,
            len(target_candidates),
        )
        target_candidates_count = len(target_candidates)
        for cand, _, households in target_candidates:
            LOGGER.info("Target complex: id=%s name=%s households=%s", cand.complex_id, cand.name, households)

        if progress_callback:
            progress_callback(
                {
                    "event": "query_target_ready",
                    "query": query,
                    "current": 0,
                    "total": target_candidates_count,
                }
            )

        for cand, _, _ in target_candidates:
            if cand.complex_id in processed_complex_ids:
                continue
            processed_complex_ids.add(cand.complex_id)
            attempted_complexes += 1

            if cand.complex_id == seed.complex_id:
                payloads = seed_payloads
            else:
                payloads = payload_cache.get(cand.complex_id)
                if payloads is None:
                    payloads = fetch_kb_complex_payloads(session, cand.complex_id)
                    payload_cache[cand.complex_id] = payloads

            df = build_dataframe_from_kb(query=query, candidate=cand, payloads=payloads)
            if df.empty:
                failed_complexes += 1
                if progress_callback:
                    progress_callback(
                        {
                            "event": "query_progress",
                            "query": query,
                            "current": attempted_complexes,
                            "total": target_candidates_count,
                            "complex_name": cand.name,
                            "success": False,
                        }
                    )
                continue
            success_complexes += 1
            crawled_info.append(cand)
            all_frames.append(df)

            main_data = payloads.get("main") if isinstance(payloads.get("main"), dict) else {}
            marker = _extract_marker_row(
                cand=cand,
                main_data=main_data,
                seed_query=query,
                is_seed=(cand.complex_id == seed.complex_id),
            )
            if marker and cand.complex_id not in marker_ids:
                marker_ids.add(cand.complex_id)
                marker_rows.append(marker)
            if progress_callback:
                progress_callback(
                    {
                        "event": "query_progress",
                        "query": query,
                        "current": attempted_complexes,
                        "total": target_candidates_count,
                        "complex_name": cand.name,
                        "success": True,
                    }
                )

        query_finished_at = datetime.now()
        metric = QueryMetric(
            query=query,
            started_at=query_started_at,
            finished_at=query_finished_at,
            elapsed_sec=round((query_finished_at - query_started_at).total_seconds(), 2),
            seed_complex_id=seed_complex_id,
            seed_complex_name=seed_complex_name,
            nearby_candidates=nearby_candidates_count,
            target_candidates=target_candidates_count,
            attempted_complexes=attempted_complexes,
            success_complexes=success_complexes,
            failed_complexes=failed_complexes,
        )
        metrics.append(metric)
        LOGGER.info(
            "Query metric: query=%s elapsed=%.2fs success=%s/%s failure_rate=%.1f%%",
            metric.query,
            metric.elapsed_sec,
            metric.success_complexes,
            metric.attempted_complexes,
            metric.failure_rate_pct,
        )

    if not all_frames:
        raise ValueError("수집 가능한 단지가 없습니다.")

    result_df = pd.concat(all_frames, ignore_index=True)
    markers_df = pd.DataFrame(
        marker_rows,
        columns=[
            "complex_id",
            "complex_name",
            "lat",
            "lng",
            "seed_query",
            "is_seed",
            "built_year",
            "households",
            "parking",
            "hall_type",
        ],
    )
    return result_df, selected_info, crawled_info, markers_df, metrics


def run(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="단지명 기반 자동 수집 후 엑셀 저장(KB fallback pipeline)")
    parser.add_argument("--query", type=str, default=None, help="단지명")
    parser.add_argument("--dong", type=str, default=None, help="우선 동명(예: 응암동)")
    parser.add_argument("--radius-m", type=float, default=500.0, help="주변 자동 수집 반경(미터)")
    parser.add_argument("--min-households", type=int, default=290, help="수집 최소 세대수")
    parser.add_argument("--log-level", type=str, default="INFO")
    args = parser.parse_args(argv)

    setup_logging(args.log_level)
    raw_query = args.query or input("단지명을 입력하세요: ").strip()
    if not raw_query:
        print("단지명이 비어 있습니다.")
        return 1
    try:
        result_df, selected_info, crawled_info, _, metrics = collect_dataset(
            raw_query=raw_query,
            radius_m=args.radius_m,
            min_households=args.min_households,
            preferred_dong=(args.dong.strip() if isinstance(args.dong, str) and args.dong.strip() else None),
        )
    except ValueError as exc:
        print(str(exc))
        return 1

    print(result_df.head(60))

    queries = split_queries(raw_query)
    save_stem = queries[0] if len(queries) == 1 else f"{queries[0]}_외{len(queries)-1}"
    out_path = save_output(result_df, query=save_stem)
    metric_path = save_query_metrics(metrics)
    print(f"\n저장 완료: {out_path}")
    print(f"메트릭 로그 저장: {metric_path}")
    for q, s in selected_info:
        print(f"[SEED][{q}] kb_complex_id: {s.complex_id}, kb_complex_name: {s.name}")
    print(f"총 수집 단지수(중복제거): {len(crawled_info)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
